import importlib  # 提供import语句
import sys
import time  # 提供延时功能
import xlrd  # excel文件读取
import os  # 系统操作库 用于打开chrome
import xlwt  # excel文件写入
from xlutils.copy import copy  # excel文件复制
from selenium.webdriver.common.by import By  # 用于获取网页中的相关元素、标签
from selenium import webdriver

importlib.reload(sys)  # 避免utf-8等编码问题的出现

# 伪装成浏览器，防止被识破
option = webdriver.ChromeOptions()
option.add_argument(
    '--user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.146 Safari/537.36"')  # 请求头
driver = webdriver.Chrome(options=option)  # 打开chrome浏览器
# 打开企查查登录页面
driver.get('https://www.qichacha.com/user_login')
time.sleep(20)  # 等待20s以用于完成手动登录操作

print("登录结束，开始爬取信息")


# 手动登录操作
def get_inc_list():
    # 从excel获取查询企业单位
    worksheet = xlrd.open_workbook(r'爬虫结果.xlsx')  # 打开已有企业名称的文件
    sheet1 = worksheet.sheet_by_name("企业信息")  # 取工作表名称为“企业信息”的工作表
    rows = sheet1.nrows  # 获取行数
    inc_list = []  # 用列表来装企业名称
    for i in range(0, rows):
        data = sheet1.cell_value(i, 0)  # 取第1列数据
        inc_list.append(data)
    print(inc_list)
    print(len(inc_list))  # 获取总共读取到的列表元素个数


inc_list = ["阿里巴巴", "字节跳动"]
inc_len = len(inc_list)

credit_list = []  # 用来装社会信用代码的空列表
capital_list = []

# 开始爬虫
for i in range(inc_len):
    txt = inc_list[i]
    time.sleep(1)
    if (i == 0):  # 如果是第一次 则直接向搜索框注入内容，不用清除搜索框中的内容。
        # 向搜索框注入文字
        driver.find_element(By.ID, 'searchKey').send_keys(txt)  # 这里的ID可以用开发者界面在搜索框那找到
        # 单击搜索按钮
        srh_btn = driver.find_element(By.XPATH,
                                      '/html/body/div[1]/div[2]/section[1]/div/div/div/div[1]/div/div/span/button')  # 这里的XPATH直接复制过来替换掉就可以
        srh_btn.click()
        time.sleep(2)
    else:
        # 清楚搜索框内容
        driver.find_element(By.ID, 'searchKey').clear()
        # 向搜索框注入下一个企业地址
        driver.find_element(By.ID, 'searchKey').send_keys(txt)
        # 找到并点击搜索按钮
    srh_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div/div[1]/div/div/div/div/span/button')
    srh_btn.click()

    time.sleep(2)
    btnn = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div[3]/div/div[2]/div/table/tr[1]/td[3]/div[2]/span/span[1]/a')
    btnn.click()
    time.sleep(2)

    try:
        # 社会码
        # /html/body/div[1]/div[2]/div[2]/div[3]/div/div[2]/div/table/tr[1]/td[3]/div/div[3]/div[1]/span[4]/span
        # //*[@id="cominfo"]/div[2]/table/tr[1]/td[2]/span/span[1]
        # //*[@id="cominfo"]/div[2]/table/tr[4]/td[6]/span/span[1]
        # /html/body/div[1]/div[2]/div[5]/div[2]/div/div[2]/section[2]/div[2]/table/tr[1]/td[2]/span/span[1]
        # /html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/div/div/div[1]/div[1]/div[1]/span/span/span/span[1]
        # /html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/div/div/div[1]/div[1]/div[3]/span/span[2]

        # 注册资本
        # /html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/div/div/div[1]/div[1]/div[3]/span/span[2]

        # 直接找到装在社会信用代码的标签
        credit_code = driver.find_element(By.XPATH,
                                          '//*[@id="cominfo"]/div[2]/table/tr[4]/td[2]/span/span[1]').text
                                          # '/html/body/div/div[2]/div[2]/div[3]/div/div[2]/div/table/tr[1]/td[3]/div/div[3]/div[1]/span[2]/span').text
                                          # '/html/body/div/div[2]/div[2]/div[3]/div/div[2]/div/table/tr[1]/td[3]/div/span/span[1]/a/span').text
                                          # '/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/div/div/div[1]/div[1]/div[3]/span/span[2]').text
                                          # '/html/body/div[1]/div[2]/div[2]/div[3]/div/div[2]/div/table/tr[1]/td[3]/div/div[3]/div[1]/span[4]/span').text
        # 字符化，防止后续出现编码问题
        credit_code = str(credit_code)
        print(credit_code, "社会码")  # 每找到一条，则打印一条在终端上。
        credit_list.append(credit_code)  # 添加到列表

        # /html/body/div/div[2]/div[2]/div[1]/div/div[2]/div/div/div/div[1]/div[1]/div[3]/span/span[2]
        # /html/body/div/div[2]/div[2]/div[1]/div/div[2]/div/div/div/div[1]/div[1]/div[3]/span/span[2]

        capital = driver.find_element(By.XPATH,
                                      '/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/div/div/div[1]/div[1]/div[3]/span/span[2]').text

        capital = str(capital)
        print(capital, "注册资本")
        capital_list.append(capital)
    except:
        credit_code = '没有找到社会码,可能是企业名称非全称导致筛选信息模糊...'
        credit_list.append(credit_code)  # 没找到也要添加到列表里面去，实现与企业一一对应
        print("没有找到社会码,可能是企业名称非全称导致筛选信息模糊...")

        capital = '没有找到注册资本,可能是企业名称非全称导致筛选信息模糊...'
        capital_list.append(capital)  # 没找到也要添加到列表里面去，实现与企业一一对应
        print("没有找到注册资本,可能是企业名称非全称导致筛选信息模糊...")

time.sleep(1000)

# 读取到的数据存入excel中
def save_to_excel():
    global i
    import openpyxl as pl
    wb = pl.load_workbook(r'爬虫结果.xlsx')  # 打开工作簿
    ws = wb.active  # 读取第一张工作表
    i = 0
    while i < len(credit_list):
        ws.cell(row=i + 1, column=2,
                value=credit_list[i])  # 将信用代码的列表中的每个数据写入到excel中，cell(row=1,column=1,values = 10) 表示A1单元格内容是10.
        print(f'{credit_list[i]}信用码正在存入{ws.cell(i + 1, 1).value}中')  # cell(1,1).value表示单元格A1的值
        i = i + 1
    wb.save(r'爬虫结果.xlsx')  # 保存文件，这一步必须需要，否则文件没保存等于白搞。
    print("Already get things done ,please check the file now!")  # 可以自行编辑程序完成之后的输出内容。
